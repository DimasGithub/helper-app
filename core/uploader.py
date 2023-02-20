import json
import asyncio
import aiohttp

import requests
import datetime
from datetime import time
from requests.structures import CaseInsensitiveDict
from openpyxl import load_workbook
from abc import ABC, abstractmethod
from django.core.files.storage import FileSystemStorage
from django.conf import settings
from django.core.files.uploadedfile import InMemoryUploadedFile
class EkinerjaException(Exception):
    def __init__(self, message):
        self.message = message

class BaseApiEkinerja(ABC):
    url_base_ekinerja ='https://api.e-kinerja.purbalinggakab.go.id'
    url_base_pegawai = 'https://api.pegawai.e-kinerja.purbalinggakab.go.id'

    @abstractmethod    
    def requests_data(self):
        pass

class InfoEmployee(BaseApiEkinerja):
    def __init__(self, nip):
        self.nip = nip

    def requests_data(self):
        endpoint = f"{self.url_base_pegawai}/pegawai/nip/{self.nip}"
        return requests.get(endpoint)

class ListDailyReport(BaseApiEkinerja):
    def __init__(self, nip):
        self.nip = nip
        
    def requests_data(self):
        endpoint = f"{self.url_base_ekinerja}/laporan-harian"
        payload = {'nip':self.nip}
        return requests.get(endpoint, params=payload)

class ListPerformAggrement(BaseApiEkinerja):
    def __init__(self, nip:str, filter:dict):
        self.nip = nip
        self.month = filter.get('month')
        self.year = filter.get('year')

    def requests_data(self):
        endpoint = f"{self.url_base_ekinerja}/perjanjian-kinerja/find/rencana/{self.nip}/{self.month}/{self.year}"
        return requests.get(endpoint)

class InfoYearlySkp(BaseApiEkinerja):
    def __init__(self, nip):
        self.nip = nip
        
    def requests_data(self):
        endpoint = f"{self.url_base_ekinerja}/skp-tahunan/{self.nip}"
        return requests.get(endpoint)

class DeleteAllData(ListDailyReport):
    def __init__(self, nip) -> None:
        self.nip = nip
    def requests_data(self) -> None:
        endpoint = f"{self.url_base_ekinerja}/laporan-harian"
        headers = CaseInsensitiveDict()
        headers["Content-Type"] = "application/json"
        list_daily_report = ListDailyReport(self.nip).requests_data()
        resp_list = json.loads(list_daily_report.text)
        list_id = []
        for resp in resp_list.get('items'):
            list_id.append(resp.get('id'))
        for id in list_id:
            resp = requests.delete(endpoint, headers = headers, data=json.dumps({'id':int(id)}, indent=4))
            print(resp)
            
class DeleteReportDaily(BaseApiEkinerja):
    def __init__(self, id) -> None:
        self.id = id
    def requests_data(self)-> None:
        endpoint = f"{self.url_base_ekinerja}/laporan-harian"
        headers = CaseInsensitiveDict()
        headers["Content-Type"] = "application/json"
        return requests.delete(endpoint, headers = headers, data=json.dumps({'id':self.id}, indent=4))

class PostingDailyReport(ListDailyReport, InfoEmployee, InfoYearlySkp, ListPerformAggrement ):

    def __init__(self, nip:str, filename:str):
        self.nip = nip
        self.filename = filename

    def xls_to_dict(self, filename)->dict:
        print(filename)
        wb = load_workbook(filename)
        worksheet =  wb.active
        list_activities = []
        for i in range(1, worksheet.max_row):
            list_activity = []
            for col in worksheet.iter_cols(1, worksheet.max_column):
                list_activity.append(col[i].value)
            if not any(elem is None for elem in list_activity): list_activities.append(list_activity)

        new_list=[]
        for a in list_activities:
            dict_activities = {}
            dict_activities['date'] = a[0]
            dict_activities['start_time'] = a[1]
            dict_activities['last_time'] = a[2]
            if a[3] in ['N','n']:
                dict_activities['perform_relation_monthly'] = False
                dict_activities['perform_monthly'] = ''
            elif a[3] in ['Y', 'y']:
                dict_activities['perform_relation_monthly'] = True
                dict_activities['perform_monthly'] = a[4]
            dict_activities['activity_summeries'] = a[5]
            dict_activities['count'] = a[6]
            new_list.append(dict_activities)
        return new_list

    def initial_data_user(self, nip)->dict:

        info_user = InfoEmployee(nip).requests_data()
        if info_user.status_code != 200:
            raise EkinerjaException(message="Info user not Found.")
        data_user = json.loads(info_user.text)
        info_yaerly_skp = InfoYearlySkp(nip).requests_data()

        if info_yaerly_skp.status_code != 200:
            raise EkinerjaException(message="Info yearly skp not found.")
        response_info_yaerly_skp = json.loads(info_yaerly_skp.text)
        data_info_yaerly_skp = response_info_yaerly_skp[0]
        intial_data_user= {
            "nip":nip,
            "foto":data_user.get('data').get('photo'),
            "nama":data_info_yaerly_skp.get('userNama'),
            "jab":data_info_yaerly_skp.get('userJab'),
            "atasan":data_info_yaerly_skp.get('atasanPejabatPenilaiNip'),
            "atasanNama":data_info_yaerly_skp.get('atasanPejabatPenilaiNama'),
            "atasanJab": data_info_yaerly_skp.get('atasanPejabatPenilaiJab'),
            "skpd":data_user.get('data').get('unit').get('id'),
            "subUnit":data_user.get('data').get('subunit').get('id')
        }
        return intial_data_user
   
    #if using duplicate validation need the time one two minutes.
    def is_duplicated(self, data:dict)->dict:
        list_daily = ListDailyReport(nip=self.nip).requests_data()
        list_data  = json.loads(list_daily.text)
        for daily in list_data.get('items'):
            duplicate_num = 0
            date = daily.get('tanggal').split()
            if date[1].lower() == 'januari':
                date[1] = 'January'
            new_date_string = ' '.join(date)  
            if datetime.datetime.strptime(new_date_string, "%d %B %Y").strftime("%Y-%m-%d") == data.get('tanggal'):
                duplicate_num += 1
            if daily.get('waktuMulai') == data.get('waktuMulai'):
                duplicate_num += 1
            if daily.get('waktuSelesai') == data.get('waktuSelesai'):
                duplicate_num += 1
            if daily.get('rencanaKinerja').get('id') == data.get('rencanaKinerjaId'):
                duplicate_num += 1
            if int(daily.get('jumlah')) == int(data.get('jumlah')):
                duplicate_num += 1
            
            if duplicate_num == 5: return True
        return False

    def requests_data(self) ->None:
        endpoint = f"{self.url_base_ekinerja}/laporan-harian"
        headers = CaseInsensitiveDict()
        headers["Content-Type"] = "application/json"
        initial_data_user = self.initial_data_user(self.nip)
        data_xls = self.xls_to_dict(self.filename)
        for item in data_xls:
            def hours_minutes_time(start_time:str, last_time:str ):
                start_time = datetime.datetime.strptime(start_time,"%H:%M:%S").time()
                last_time = datetime.datetime.strptime(last_time,"%H:%M:%S").time()
                start_time = datetime.datetime.combine(datetime.datetime.today(), start_time)
                last_time = datetime.datetime.combine(datetime.datetime.today(), last_time)
                different_time = last_time - start_time
                minutes, _ = divmod(different_time.seconds, 60)
                hours, minutes = divmod(minutes, 60)
                return (int(hours), int(minutes))

            hours, minutes = hours_minutes_time(item.get('start_time'), item.get('last_time'))
            waktu = f"{hours} Jam {minutes} Menit" if hours > 0 else f"{minutes} Menit"
            data ={
                "tanggal":item.get('date'),
                "waktuMulai":item.get('start_time'),
                "waktuSelesai":item.get('last_time'),
                "jam":hours,
                "menit":minutes,
                "waktu":waktu,
                "jumlah":item.get('count'),
                "uraian":item.get('activity_summeries'),
                "skp":'ada' if item.get('perform_relation_monthly') else 'tidak-ada',
                'rencanaKinerjaId': item.get('perform_monthly')
                }
            
            def get_perform_aggrement(nip:str, rencanaKinerjaId:str)->dict:
                list_perform_aggr = ListPerformAggrement(nip, filter={'month':'Januari', 'year':'2023'}).requests_data()
                if list_perform_aggr.status_code != 200:
                    raise EkinerjaException(message="List Perform Aggrement Not Found.")

                response_perform_aggr = json.loads(list_perform_aggr.text)

                data_perform_aggrs = response_perform_aggr.get('data').get('rencanaKinerja')
                for data_perform_aggr in data_perform_aggrs:
                    if " ".join(rencanaKinerjaId.lower().split()) == " ".join(data_perform_aggr.get('targetSkp').get('kegiatanTahunan').lower().split()):
                        return data_perform_aggr
            
            perform_aggr = get_perform_aggrement(self.nip, data.get('rencanaKinerjaId'))
            data_relation_perform = {
                "rencanaKinerjaId": perform_aggr.get('id'),
                "output": perform_aggr.get('satuanKuantitas'),
            }
            data.update(data_relation_perform)
            data.update(initial_data_user)
            if not self.is_duplicated(data=data):
                data = json.dumps(data, indent=4)
                response = requests.post(endpoint, headers = headers,data=data)

if __name__ == "__main__":
    # PostingDailyReport('199510292022031004', filename='core/format_kinerja2.xlsx').requests_data()
    DeleteAllData('199510292022031004').requests_data()